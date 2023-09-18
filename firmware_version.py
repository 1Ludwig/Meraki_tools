#!/usr/bin/python3

import api_info
import meraki
from datetime import datetime
from openpyxl import Workbook
from tqdm import tqdm

API_KEY = api_info.api_nyckel
dashboard = meraki.DashboardAPI(API_KEY, suppress_logging=True)

"""Excel info"""
wb = Workbook()
ws = wb.active
ws["A1"] = "Customer"
ws["B1"] = "Network"
ws["C1"] = "Device"
ws["D1"] = "Software Version"
ws["E1"] = "Device Model"
ws["F1"] = "Serial Number"
ws["G1"] = "Tags"


def main():
    """Fetch organisation ID's and put them in a dict"""
    org_dict = dashboard.organizations.getOrganizations()

    org_list = []
    network_list_id = []

    """From the dict of organisations, make a list of networks associated with the organisation"""
    for org_list_loop in org_dict:
        org_list.append(org_list_loop.get("id"))

        try:
            """If a network doesn't have a License like a lab or test it gives API Error then just skip it"""
            network_dict = dashboard.organizations.getOrganizationNetworks(
                org_list_loop.get("id"), suppress_logging=True
            )
        except meraki.APIError:
            # with open("failed_orgs.txt", "w+"):
            #    print(f"Failed to run script on org: {org_list_loop}" + "\n").encode("utf-8")
            continue

        """From the list of Networks, make a list of all associated devices"""
        n = org_list_loop.get("name")
        for network_list_loop in tqdm(network_dict, desc=str(f"{n}")):
            network_list_id.append(network_list_loop.get("id"))
            device_list = dashboard.networks.getNetworkDevices(
                network_list_loop.get("id")
            )

            """From the list of devices, loop through them all and fetch information to print"""
            for device in device_list:
                name = str(device.get("name"))
                software_version = str(device.get("firmware"))
                model = str(device.get("model"))
                serial_number = str(device.get("serial"))
                customer = str(org_list_loop.get("name"))
                network_name = str(network_list_loop.get("name"))
                tags = ";".join([str(item) for item in device.get("tags")])

                """Write to Excel Sheet"""
                exel_info = (
                    customer,
                    network_name,
                    name,
                    software_version,
                    model,
                    serial_number,
                    tags,
                )
                ws.append(exel_info)
                wb.save("firmware_report.xlsx")


if __name__ == "__main__":
    start_time = datetime.now()
    main()
    wb.save("firmware_report.xlsx")
    print("\nElapsed time: " + str(datetime.now() - start_time))
